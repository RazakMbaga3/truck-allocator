"""Find the PO quantity column in the RM PO sheets."""
import openpyxl
from pathlib import Path

BASE = Path(__file__).parent.parent
path = BASE / "Raw Material -Transporter Mater, Supplier Master & RM Purchase-25.04.xlsx"
wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

ws = wb["RM Purchase Order-Coal"]
# Print header row (row 5) fully
print("HEADER ROW (row 5):")
for i, row in enumerate(ws.iter_rows(min_row=5, max_row=5, values_only=True)):
    for j, v in enumerate(row):
        if v is not None:
            print(f"  col {j}: {v!r}")

print("\nFULL DATA ROW (row 6):")
for i, row in enumerate(ws.iter_rows(min_row=6, max_row=6, values_only=True)):
    for j, v in enumerate(row):
        if v is not None:
            print(f"  col {j}: {v!r}")
wb.close()
