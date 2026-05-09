"""
scripts/import_lcl_data.py — Import Lake Cement data into local Odoo 15.

Creates in order:
  1. Products  — RM materials (Coal, Clinker, Gypsum, Iron Ore) + cement products
  2. Vendors   — RM suppliers from Raw Material Vendor Master-Supp sheet
  3. Customers — from Customer master.xlsx
  4. POs       — all RM Purchase Orders (Coal, Clinker, Gypsum, Iron Ore)
  5. SOs       — most recent 300 approved Sale Orders

Safe to re-run — checks for existing records before creating.

Usage:
    python scripts/import_lcl_data.py
    python scripts/import_lcl_data.py --limit-so 100   # fewer SOs
    python scripts/import_lcl_data.py --skip-customers  # skip slow customer import
"""

from __future__ import annotations

import argparse
import sys
import xmlrpc.client
from datetime import datetime
from pathlib import Path

import openpyxl

BASE = Path(__file__).parent.parent

URL  = "http://localhost:8069"
DB   = "Razak"
USER = "digital@lakecement.co.tz"
PWD  = "Mbagarazack617@"


# ── Odoo helpers ──────────────────────────────────────────────────────────────

def connect():
    common = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/common")
    uid = common.authenticate(DB, USER, PWD, {})
    if not uid:
        print("ERROR: Authentication failed")
        sys.exit(1)
    models = xmlrpc.client.ServerProxy(f"{URL}/xmlrpc/2/object")
    return uid, models


def search(models, uid, model, domain, fields=None, limit=None):
    kw = {"fields": fields or ["id"]}
    if limit:
        kw["limit"] = limit
    return models.execute_kw(DB, uid, PWD, model, "search_read", [domain], kw)


def create(models, uid, model, vals, ctx=None):
    kw = {}
    if ctx:
        kw["context"] = ctx
    return models.execute_kw(DB, uid, PWD, model, "create", [vals], kw)


def call(models, uid, model, method, ids, kw=None):
    return models.execute_kw(DB, uid, PWD, model, method, [ids], kw or {})


def find_or_create(models, uid, model, domain, vals, label=""):
    existing = search(models, uid, model, domain, ["id", "name"])
    if existing:
        return existing[0]["id"], False
    new_id = create(models, uid, model, vals)
    return new_id, True


# ── Step 1: Products ──────────────────────────────────────────────────────────

RM_PRODUCTS = [
    {"name": "COAL",        "default_code": "RM000001", "categ": "Raw Material"},
    {"name": "GYPSUM",      "default_code": "RM000003", "categ": "Raw Material"},
    {"name": "IRON ORE",    "default_code": "RM000004", "categ": "Raw Material"},
    {"name": "CLINKER SPG", "default_code": "RM000014", "categ": "Raw Material"},
]

CEMENT_PRODUCTS = [
    {"name": "CEM II A-L 42.5 R", "default_code": "CEM42R",  "categ": "Finished Product"},
    {"name": "CEM II B-M 42.5 N", "default_code": "CEM42N",  "categ": "Finished Product"},
]


def setup_products(models, uid):
    print("\n[1/5] PRODUCTS")

    # Find an existing weight-class UoM (t, Tonne, MT, kg — whatever exists)
    mt_uom_id = None
    for uom_name in ["t", "MT", "Tonne", "Tonnes", "ton", "kg"]:
        results = search(models, uid, "uom.uom", [["name", "=", uom_name]], ["id", "name"])
        if results:
            mt_uom_id = results[0]["id"]
            print(f"  Using UoM: {results[0]['name']} (id={mt_uom_id})")
            break
    if not mt_uom_id:
        # Fall back to the default unit (id=1 in standard Odoo)
        results = search(models, uid, "uom.uom", [], ["id", "name"], limit=1)
        mt_uom_id = results[0]["id"] if results else 1
        print(f"  Using fallback UoM id={mt_uom_id}")

    # Get or create product category
    def get_or_create_categ(name):
        cid, created = find_or_create(
            models, uid, "product.category",
            [["name", "=", name]],
            {"name": name},
        )
        if created:
            print(f"  Created category: {name}")
        return cid

    rm_categ_id = get_or_create_categ("Raw Material")
    fg_categ_id = get_or_create_categ("Finished Product")

    product_ids = {}
    for p in RM_PRODUCTS + CEMENT_PRODUCTS:
        categ_id_for_p = rm_categ_id if p["categ"] == "Raw Material" else fg_categ_id
        pid, created = find_or_create(
            models, uid, "product.template",
            [["default_code", "=", p["default_code"]]],
            {
                "name": p["name"],
                "default_code": p["default_code"],
                "type": "product",
                "categ_id": categ_id_for_p,
                "uom_id": mt_uom_id,
                "uom_po_id": mt_uom_id,
            },
        )
        status = "Created" if created else "Exists"
        print(f"  {status}: {p['name']} [{p['default_code']}]")
        # Get product.product (variant) id
        variants = search(models, uid, "product.product",
                          [["product_tmpl_id", "=", pid]], ["id", "default_code"])
        product_ids[p["default_code"]] = variants[0]["id"] if variants else None

    return product_ids, mt_uom_id


# ── Step 2: Vendors ───────────────────────────────────────────────────────────

def import_vendors(models, uid):
    print("\n[2/5] VENDORS (RM Suppliers)")
    path = BASE / "Raw Material -Transporter Mater, Supplier Master & RM Purchase-25.04.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Raw Material Vendor Master-Supp"]

    vendor_map = {}  # code → odoo partner id
    created_count = 0
    exists_count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        code = row[0]
        name = row[1]
        if not code or not name or str(code).startswith("Code"):
            continue
        code = str(code).strip()
        name = str(name).strip()

        region = str(row[5]).strip() if row[5] else ""
        phone  = str(row[6]).strip() if row[6] else ""
        email  = str(row[8]).strip() if row[8] else ""

        pid, created = find_or_create(
            models, uid, "res.partner",
            [["ref", "=", code]],
            {
                "name": name,
                "ref": code,
                "supplier_rank": 1,
                "customer_rank": 0,
                "company_type": "company",
                "city": region.replace(" (TZ)", ""),
                "country_id": _tz_country_id(models, uid),
                "phone": phone[:20] if phone else False,
                "email": email[:100] if email else False,
                "comment": f"RM Supplier — imported from LCL data",
            },
        )
        vendor_map[code] = pid
        if created:
            created_count += 1
        else:
            exists_count += 1

    wb.close()
    print(f"  Created: {created_count}  |  Already existed: {exists_count}")
    return vendor_map


# ── Step 3: Customers ─────────────────────────────────────────────────────────

def import_customers(models, uid, limit=None):
    print(f"\n[3/5] CUSTOMERS (limit={limit or 'all'})")
    path = BASE / "Customer master.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    customer_map = {}  # code → odoo partner id
    created_count = 0
    exists_count = 0
    count = 0

    tz_id = _tz_country_id(models, uid)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if limit and count >= limit:
            break
        code = row[0]
        name = row[1]
        if not code or not name:
            continue
        code = str(code).strip()
        name = str(name).strip()

        city     = str(row[10]).strip() if row[10] else ""
        region   = str(row[12]).strip() if row[12] else ""
        district = str(row[13]).strip() if row[13] else ""
        phone    = str(row[18]).strip() if row[18] else ""
        email    = str(row[19]).strip() if row[19] else ""

        pid, created = find_or_create(
            models, uid, "res.partner",
            [["ref", "=", code]],
            {
                "name": name,
                "ref": code,
                "customer_rank": 1,
                "supplier_rank": 0,
                "company_type": "company",
                "city": city,
                "country_id": tz_id,
                "phone": phone[:20] if phone else False,
                "email": email[:100] if email else False,
            },
        )
        customer_map[code] = pid
        if created:
            created_count += 1
        else:
            exists_count += 1
        count += 1

        if count % 100 == 0:
            print(f"  ... {count} customers processed")

    wb.close()
    print(f"  Created: {created_count}  |  Already existed: {exists_count}")
    return customer_map


# ── Step 4: Purchase Orders ───────────────────────────────────────────────────

RM_ITEM_CODE_TO_PRODUCT = {
    "RM000001": "RM000001",   # COAL
    "RM000003": "RM000003",   # GYPSUM
    "RM000004": "RM000004",   # IRON ORE
    "RM000014": "RM000014",   # CLINKER SPG
}

PO_SHEETS = {
    "RM Purchase Order-Coal":     "RM000001",
    "RM Purchase Order-Clinker":  "RM000014",
    "RM Purchase Order-Gypsum":   "RM000003",
    "RM Purchase Order-Iron ore ": "RM000004",
}


def import_purchase_orders(models, uid, vendor_map, product_ids, mt_uom_id):
    print("\n[4/5] PURCHASE ORDERS")
    path = BASE / "Raw Material -Transporter Mater, Supplier Master & RM Purchase-25.04.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)

    # Get the user's own company (safe for multi-company)
    user_data = search(models, uid, "res.users", [["id", "=", uid]], ["company_id"])
    company_id = user_data[0]["company_id"][0] if user_data and user_data[0].get("company_id") else 1

    created_count = 0
    skipped_count = 0
    error_count = 0

    for sheet_name, default_item_code in PO_SHEETS.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Sheet not found: {sheet_name}")
            continue

        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=6, values_only=True):
            seq = row[0]
            if seq is None or not isinstance(seq, (int, float)):
                continue

            vendor_code = str(row[10]).strip() if row[10] else None
            vendor_name = str(row[11]).strip() if row[11] else "Unknown"
            po_name     = str(row[15]).strip() if row[15] else None
            item_code   = str(row[20]).strip() if row[20] else default_item_code
            qty         = float(row[22]) if row[22] else 0.0
            po_date_str = str(row[16]).strip() if row[16] else None

            if not po_name or not vendor_code or qty <= 0:
                continue

            # Check if PO already exists
            existing = search(models, uid, "purchase.order",
                              [["partner_ref", "=", po_name]], ["id"])
            if existing:
                skipped_count += 1
                continue

            # Find or create vendor
            if vendor_code in vendor_map:
                partner_id = vendor_map[vendor_code]
            else:
                vp, _ = find_or_create(
                    models, uid, "res.partner",
                    [["ref", "=", vendor_code]],
                    {"name": vendor_name, "ref": vendor_code,
                     "supplier_rank": 1, "company_type": "company",
                     "country_id": _tz_country_id(models, uid)},
                )
                partner_id = vp
                vendor_map[vendor_code] = vp

            # Get product id
            product_id = product_ids.get(item_code)
            if not product_id:
                error_count += 1
                continue

            # Parse date
            po_date = None
            if po_date_str:
                for fmt in ["%d-%m-%Y", "%Y-%m-%d"]:
                    try:
                        po_date = datetime.strptime(po_date_str, fmt).strftime("%Y-%m-%d %H:%M:%S")
                        break
                    except Exception:
                        pass

            try:
                po_vals = {
                    "partner_id": partner_id,
                    "partner_ref": po_name,
                    "date_order": po_date or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "order_line": [(0, 0, {
                        "product_id": product_id,
                        "product_qty": qty,
                        "product_uom": mt_uom_id,
                        "price_unit": 0.0,
                        "name": f"{item_code} — {po_name}",
                        "date_planned": po_date or datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    })],
                }
                ctx = {"allowed_company_ids": [company_id], "company_id": company_id}
                po_id = create(models, uid, "purchase.order", po_vals, ctx=ctx)
                # Confirm the PO
                call(models, uid, "purchase.order", "button_confirm", [po_id])
                created_count += 1
                print(f"  Created PO: {po_name} | {vendor_name} | {item_code} | {qty} MT")
            except Exception as e:
                error_count += 1
                print(f"  ERROR creating {po_name}: {e}")

    wb.close()
    print(f"  Created: {created_count}  |  Skipped (exists): {skipped_count}  |  Errors: {error_count}")


# ── Step 5: Sale Orders ───────────────────────────────────────────────────────

def import_sale_orders(models, uid, customer_map, product_ids, mt_uom_id, limit=300):
    print(f"\n[5/5] SALE ORDERS (most recent {limit})")
    path = BASE / "approved sales orders 1st April'25 to 24th April'26.xlsx"
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb.active

    # Get pricelist
    pricelists = search(models, uid, "product.pricelist", [], ["id"], limit=1)
    pricelist_id = pricelists[0]["id"] if pricelists else None

    # Use the user's own company (avoids multi-company access errors)
    user_data = search(models, uid, "res.users", [["id", "=", uid]], ["company_id"])
    company_id = user_data[0]["company_id"][0] if user_data and user_data[0].get("company_id") else 1

    tz_id = _tz_country_id(models, uid)
    created_count = 0
    skipped_count = 0
    error_count = 0
    count = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if count >= limit:
            break

        customer_raw  = row[0]   # "[CUS909] Name"
        so_ref        = row[1]   # SO/2026/01246
        order_date    = row[2]   # datetime
        product_name  = row[9]   # "CEM II A-L 42.5 R"
        qty_mt        = row[11]  # numeric
        destination   = row[12]  # "TEMEKE"
        transporter   = row[13]  # "CUSTOMER ARRANGED"
        district      = row[14]
        region        = row[15]

        if not so_ref or not customer_raw or not qty_mt:
            continue

        # Parse customer code
        customer_code = None
        customer_name = str(customer_raw)
        if str(customer_raw).startswith("["):
            bracket_end = str(customer_raw).find("]")
            if bracket_end > 0:
                customer_code = str(customer_raw)[1:bracket_end]
                customer_name = str(customer_raw)[bracket_end + 2:].strip()

        # Check if SO already exists
        existing = search(models, uid, "sale.order",
                          [["client_order_ref", "=", so_ref]], ["id"])
        if existing:
            skipped_count += 1
            count += 1
            continue

        # Find or create customer
        if customer_code and customer_code in customer_map:
            partner_id = customer_map[customer_code]
        else:
            search_domain = [["ref", "=", customer_code]] if customer_code else [["name", "=", customer_name]]
            existing_p = search(models, uid, "res.partner", search_domain, ["id"])
            if existing_p:
                partner_id = existing_p[0]["id"]
                if customer_code:
                    customer_map[customer_code] = partner_id
            else:
                partner_id = create(models, uid, "res.partner", {
                    "name": customer_name,
                    "ref": customer_code or "",
                    "customer_rank": 1,
                    "company_type": "company",
                    "city": str(destination).strip() if destination else "",
                    "country_id": tz_id,
                })
                if customer_code:
                    customer_map[customer_code] = partner_id

        # Map product
        if product_name and "42.5 R" in str(product_name):
            product_id = product_ids.get("CEM42R")
        elif product_name and "42.5 N" in str(product_name):
            product_id = product_ids.get("CEM42N")
        else:
            product_id = product_ids.get("CEM42R")  # default

        if not product_id:
            error_count += 1
            count += 1
            continue

        # Format date
        if isinstance(order_date, datetime):
            date_str = order_date.strftime("%Y-%m-%d %H:%M:%S")
        else:
            date_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        try:
            so_vals = {
                "partner_id": partner_id,
                "client_order_ref": so_ref,
                "date_order": date_str,
                "note": f"Dest: {destination} | Transporter: {transporter} | District: {district} | Region: {region}",
                "order_line": [(0, 0, {
                    "product_id": product_id,
                    "product_uom_qty": float(qty_mt) * 20,  # MT → 50kg bags
                    "product_uom": mt_uom_id,
                    "price_unit": 0.0,
                    "name": f"{product_name or 'Cement'} → {destination}",
                })],
            }
            if pricelist_id:
                so_vals["pricelist_id"] = pricelist_id

            so_ctx = {"allowed_company_ids": [company_id], "company_id": company_id}
            so_id = create(models, uid, "sale.order", so_vals, ctx=so_ctx)
            call(models, uid, "sale.order", "action_confirm", [so_id])
            created_count += 1
            count += 1

            if created_count % 20 == 0:
                print(f"  ... {created_count} SOs created so far")
        except Exception as e:
            error_count += 1
            count += 1
            print(f"  ERROR creating {so_ref}: {e}")

    wb.close()
    print(f"  Created: {created_count}  |  Skipped (exists): {skipped_count}  |  Errors: {error_count}")


# ── Utility ───────────────────────────────────────────────────────────────────

_tz_country_cache = None

def _tz_country_id(models, uid):
    global _tz_country_cache
    if _tz_country_cache is None:
        results = search(models, uid, "res.country", [["code", "=", "TZ"]], ["id"])
        _tz_country_cache = results[0]["id"] if results else False
    return _tz_country_cache


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Import LCL data into local Odoo 15")
    parser.add_argument("--limit-so", type=int, default=300, help="Max SOs to import (default 300)")
    parser.add_argument("--limit-customers", type=int, default=None, help="Max customers to import")
    parser.add_argument("--skip-customers", action="store_true", help="Skip customer import")
    parser.add_argument("--skip-so", action="store_true", help="Skip sale order import")
    parser.add_argument("--skip-po", action="store_true", help="Skip purchase order import")
    args = parser.parse_args()

    print("=" * 60)
    print("  NYATI CEMENT — LCL Data Import")
    print(f"  Target: {URL}  DB: {DB}")
    print("=" * 60)

    uid, models = connect()
    print(f"  Authenticated as uid={uid}")

    product_ids, mt_uom_id = setup_products(models, uid)
    vendor_map = import_vendors(models, uid)

    customer_map = {}
    if not args.skip_customers:
        customer_map = import_customers(models, uid, limit=args.limit_customers)
    else:
        print("\n[3/5] CUSTOMERS — skipped")

    if not args.skip_po:
        import_purchase_orders(models, uid, vendor_map, product_ids, mt_uom_id)
    else:
        print("\n[4/5] PURCHASE ORDERS — skipped")

    if not args.skip_so:
        import_sale_orders(models, uid, customer_map, product_ids, mt_uom_id, limit=args.limit_so)
    else:
        print("\n[5/5] SALE ORDERS — skipped")

    print("\n" + "=" * 60)
    print("  Import complete. Restart the FastAPI app to sync.")
    print("=" * 60)


if __name__ == "__main__":
    main()
