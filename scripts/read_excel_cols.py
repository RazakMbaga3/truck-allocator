"""Read column headers and sample rows from the key Excel files."""
import sys
from pathlib import Path
import openpyxl

BASE = Path(__file__).parent.parent

files = {
    "SO": BASE / "approved sales orders 1st April'25 to 24th April'26.xlsx",
    "RM_PO": BASE / "Raw Material -Transporter Mater, Supplier Master & RM Purchase-25.04.xlsx",
    "CUSTOMER": BASE / "Customer master.xlsx",
}

for label, path in files.items():
    print("=" * 70)
    print(f"{label}: {path.name}")
    print("=" * 70)
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
        if not rows:
            continue
        print(f"\n  Sheet: '{sheet_name}'")
        print(f"  Headers: {rows[0]}")
        if len(rows) > 1:
            print(f"  Row 2:   {rows[1]}")
        if len(rows) > 2:
            print(f"  Row 3:   {rows[2]}")
    wb.close()
    print()
