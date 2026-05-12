"""
app/services/excel_export.py — Excel export with Nyati branding.

Generates formatted Excel reports showing:
  - Allocated trucks (rows)
  - Orders assigned to each truck (sub-rows)
  - Nyati Cement branding throughout
"""

from datetime import datetime
from io import BytesIO

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# Nyati brand colors
NYATI_NAVY = "173158"        # Deep navy
NYATI_ORANGE = "F49545"      # Brand orange
NYATI_GREEN = "239557"       # Brand green
NYATI_LIGHT_GRAY = "F5F5F5"  # Light gray for alternating rows
WHITE = "FFFFFF"
DARK_TEXT = "1F1F1F"

# Borders
THIN_BORDER = Border(
    left=Side(style="thin", color="D3D3D3"),
    right=Side(style="thin", color="D3D3D3"),
    top=Side(style="thin", color="D3D3D3"),
    bottom=Side(style="thin", color="D3D3D3"),
)

THICK_BORDER = Border(
    left=Side(style="medium", color=NYATI_NAVY),
    right=Side(style="medium", color=NYATI_NAVY),
    top=Side(style="medium", color=NYATI_NAVY),
    bottom=Side(style="medium", color=NYATI_NAVY),
)


def generate_final_status_report(allocations: list) -> BytesIO:
    """
    Generate Excel report of final status allocations.
    
    Args:
        allocations: List of TruckAllocation objects with:
          - schedule (TruckSchedule with truck_plate, driver_name, etc.)
          - items (list of AllocationItem)
          - status, remarks, released_at, loaded_at
    
    Returns:
        BytesIO: Excel file data ready to send as response
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Status"
    
    # Column widths
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 20
    
    # ─── HEADER ──────────────────────────────────────────────────────────────
    # Title with Nyati branding
    ws["A1"] = "LAKE CEMENT LIMITED — NYATI"
    ws["A1"].font = Font(name="Barlow Condensed", size=14, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill(start_color=NYATI_NAVY, end_color=NYATI_NAVY, fill_type="solid")
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 24
    
    # Subtitle
    ws["A2"] = "Return Truck Allocation — Final Status Report"
    ws["A2"].font = Font(name="Nunito Sans", size=11, bold=True, color=NYATI_NAVY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A2:K2")
    ws.row_dimensions[2].height = 18
    
    # Report date
    report_date = datetime.now().strftime("%d %b %Y — %H:%M")
    ws["A3"] = f"Report Generated: {report_date}"
    ws["A3"].font = Font(name="Nunito Sans", size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A3:K3")
    ws.row_dimensions[3].height = 14
    
    # Empty row
    ws.row_dimensions[4].height = 4
    
    # ─── HEADERS ROW ─────────────────────────────────────────────────────────
    headers = [
        "PO Ref",
        "Truck No",
        "Transporter",
        "Driver",
        "Origin",
        "Orders",
        "Total MT",
        "Status",
        "Ready At",
        "Loaded At",
        "Remarks"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=5, column=col)
        cell.value = header
        cell.font = Font(name="Barlow Condensed", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill(start_color=NYATI_ORANGE, end_color=NYATI_ORANGE, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[5].height = 20
    
    # ─── TRUCK DATA ──────────────────────────────────────────────────────────
    current_row = 6
    
    for alloc_idx, alloc in enumerate(allocations):
        schedule = alloc.schedule
        is_even_truck = alloc_idx % 2 == 0
        truck_fill = PatternFill(
            start_color=NYATI_LIGHT_GRAY,
            end_color=NYATI_LIGHT_GRAY,
            fill_type="solid"
        ) if is_even_truck else PatternFill()
        
        # Format times
        ready_at = ""
        loaded_at = ""
        if alloc.released_at:
            ready_at = alloc.released_at.strftime("%d %b %y")
        if alloc.loaded_at:
            loaded_at = alloc.loaded_at.strftime("%d %b %y")
        
        status_text = alloc.status.replace("_", " ")
        
        # Truck summary row
        truck_data = [
            schedule.odoo_po_name or "",
            schedule.truck_plate or "",
            schedule.transporter_name or "",
            schedule.driver_name or "",
            schedule.origin_region or "",
            len(alloc.items),
            alloc.total_tonnes,
            status_text,
            ready_at,
            loaded_at,
            alloc.remarks or "",
        ]
        
        for col, value in enumerate(truck_data, start=1):
            cell = ws.cell(row=current_row, column=col)
            cell.value = value
            cell.fill = truck_fill
            cell.border = THIN_BORDER
            cell.font = Font(name="Nunito Sans", size=9)
            
            # Format numbers
            if col == 7:  # Total MT
                cell.value = float(value)
                cell.number_format = "0.0"
                cell.alignment = Alignment(horizontal="right")
            elif col == 6:  # Orders count
                cell.alignment = Alignment(horizontal="center")
            elif col in (9, 10):  # Dates
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = Alignment(horizontal="left", wrap_text=True)
            
            # Status coloring
            if col == 8:  # Status column
                if value == "LOADED":
                    cell.font = Font(name="Nunito Sans", size=9, bold=True, color=NYATI_GREEN)
                elif value in ("WAITING LOADING", "RELEASED"):
                    cell.font = Font(name="Nunito Sans", size=9, bold=True, color=NYATI_ORANGE)
        
        ws.row_dimensions[current_row].height = 16
        current_row += 1
        
        # ─── ORDER SUB-ROWS ──────────────────────────────────────────────────
        if alloc.items:
            for item_idx, item in enumerate(alloc.items):
                item_row = current_row
                
                # Alternating row color for items
                item_fill = PatternFill(
                    start_color="FAFAFA",
                    end_color="FAFAFA",
                    fill_type="solid"
                )
                
                # Item data: Customer | Order Ref | Product | Qty | Location | Region
                item_data = [
                    "",  # A: blank (under PO)
                    f"↳ Order {item_idx + 1}",  # B: order label
                    item.customer_name or "",  # C: Transporter → Customer
                    item.order_ref or "",      # D: Driver → Order Ref
                    item.product or "",        # E: Origin → Product
                    item.quantity_tonnes,      # F: Orders → Qty
                    item.destination_location or "",  # G: Total MT → Location
                    item.region or "",         # H: Status → Region
                    "",  # I: Ready At (blank)
                    "",  # J: Loaded At (blank)
                    "",  # K: Remarks (blank)
                ]
                
                for col, value in enumerate(item_data, start=1):
                    cell = ws.cell(row=item_row, column=col)
                    cell.value = value
                    cell.fill = item_fill
                    cell.border = THIN_BORDER
                    cell.font = Font(name="Nunito Sans", size=8, italic=True, color="555555")
                    
                    if col == 6:  # Qty
                        cell.value = float(value)
                        cell.number_format = "0.0"
                        cell.alignment = Alignment(horizontal="right")
                    elif col == 2:  # Order label
                        cell.font = Font(name="Nunito Sans", size=8, italic=True, bold=True, color="555555")
                    else:
                        cell.alignment = Alignment(horizontal="left", wrap_text=True)
                
                ws.row_dimensions[item_row].height = 14
                current_row += 1
    
    # ─── FOOTER ──────────────────────────────────────────────────────────────
    current_row += 1
    ws[f"A{current_row}"] = "Total Allocations:"
    ws[f"A{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_NAVY)
    ws[f"B{current_row}"] = len(allocations)
    ws[f"B{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_ORANGE)
    
    current_row += 1
    total_mt = sum(alloc.total_tonnes for alloc in allocations)
    ws[f"A{current_row}"] = "Total Cement (MT):"
    ws[f"A{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_NAVY)
    ws[f"B{current_row}"] = total_mt
    ws[f"B{current_row}"].number_format = "0.0"
    ws[f"B{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_ORANGE)
    
    current_row += 1
    loaded_count = sum(1 for a in allocations if a.status == "LOADED")
    ws[f"A{current_row}"] = "Loaded Trucks:"
    ws[f"A{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_NAVY)
    ws[f"B{current_row}"] = loaded_count
    ws[f"B{current_row}"].font = Font(name="Barlow Condensed", size=10, bold=True, color=NYATI_GREEN)
    
    # ─── FREEZE PANES ────────────────────────────────────────────────────────
    ws.freeze_panes = "A6"
    
    # ─── EXPORT TO BYTES ─────────────────────────────────────────────────────
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
