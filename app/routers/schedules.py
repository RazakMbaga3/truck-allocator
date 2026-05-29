"""
app/routers/schedules.py — TruckSchedule API endpoints.

GET  /api/schedules              — Live list (default: available only)
GET  /api/schedules/{id}         — Schedule detail
PATCH /api/schedules/{id}/confirm-details  — Add truck plate + driver
PATCH /api/schedules/{id}/arrived          — Mark physical arrival
PATCH /api/schedules/{id}/dispatch         — Mark loaded + dispatched
POST  /api/schedules/{id}/rematch          — Force re-run matching
GET  /api/schedules/feed                   — SSE live stream
"""

from __future__ import annotations

import asyncio
import json
from datetime import datetime, timezone
from io import BytesIO
from typing import AsyncGenerator

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, delete
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import (
    AllocationProposal,
    AllocationStatus,
    TruckSchedule,
    TruckScheduleStatus,
)
from app.models.truck_schedule import TruckScheduleStatus as TSS
from app.schemas.truck_schedule import (
    TruckScheduleDetail,
    TruckScheduleListItem,
    TruckScheduleRead,
)

router = APIRouter(prefix="/api/schedules", tags=["schedules"])

# Global SSE client registry {client_id: asyncio.Queue}
_sse_clients: dict[str, asyncio.Queue] = {}


def broadcast_sse(event_type: str, data: dict) -> None:
    """Push an SSE event to all connected dashboard clients."""
    payload = json.dumps({"type": event_type, **data})
    dead = []
    for cid, q in _sse_clients.items():
        try:
            q.put_nowait(payload)
        except asyncio.QueueFull:
            dead.append(cid)
    for cid in dead:
        _sse_clients.pop(cid, None)


# ── GET /api/schedules ────────────────────────────────────────────────────────

@router.get("", response_model=list[TruckScheduleListItem])
async def list_schedules(
    status: str = Query("available", description="available | all | expected | arrived | dispatched"),
    db: AsyncSession = Depends(get_db),
):
    """
    List TruckSchedules.
    - status=available  → EXPECTED + PRE_CONFIRMED, not yet confirmed  [default]
    - status=all        → everything
    - status=expected   → only EXPECTED
    """
    q = select(TruckSchedule)

    if status == "available":
        q = q.where(
            TruckSchedule.status.in_([TSS.EXPECTED, TSS.PRE_CONFIRMED]),
            TruckSchedule.allocation_status.notin_([
                AllocationStatus.WAITING_LOADING,
                AllocationStatus.RELEASED,
                AllocationStatus.LOADED,
            ]),
        )
    elif status == "expected":
        q = q.where(TruckSchedule.status == TSS.EXPECTED)
    elif status != "all":
        q = q.where(TruckSchedule.status == status.upper())

    q = q.order_by(TruckSchedule.expected_arrival_dt.asc().nulls_last())
    result = await db.execute(q)
    schedules = result.scalars().all()
    return [TruckScheduleListItem.model_validate(s) for s in schedules]


# ── GET /api/schedules/export ────────────────────────────────────────────────

@router.get("/export")
async def export_schedules_excel(
    status: str = Query("all", description="available | all | expected | arrived | dispatched"),
    db: AsyncSession = Depends(get_db),
):
    """Export scheduled trucks as an Excel workbook."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    q = select(TruckSchedule)

    if status == "available":
        q = q.where(
            TruckSchedule.status.in_([TSS.EXPECTED, TSS.PRE_CONFIRMED]),
            TruckSchedule.allocation_status.notin_([
                AllocationStatus.RELEASED,
                AllocationStatus.LOADED,
            ]),
        )
    elif status == "expected":
        q = q.where(TruckSchedule.status == TSS.EXPECTED)
    elif status != "all":
        q = q.where(TruckSchedule.status == status.upper())

    q = q.order_by(TruckSchedule.expected_arrival_dt.asc().nulls_last())
    result = await db.execute(q)
    schedules = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Scheduled Trucks"

    headers = [
        "Upload Date",
        "Dispatch Date",
        "PO Ref",
        "Material",
        "Transporter",
        "Driver Name",
        "Phone",
        "License No.",
        "Truck No.",
        "Location",
        "Schedule Ref",
        "Truck Status",
        "Allocation Status",
        "Estimated Qty (T)",
        "Corridor",
        "Notes",
    ]
    ws.append(headers)

    for schedule in schedules:
        ws.append([
            schedule.upload_date,
            schedule.dispatch_date,
            schedule.odoo_po_name,
            schedule.raw_material_type,
            schedule.transporter_name,
            schedule.driver_name,
            schedule.driver_phone,
            schedule.driver_license_no,
            schedule.truck_plate,
            schedule.origin_region,
            schedule.schedule_ref,
            schedule.status,
            schedule.allocation_status,
            schedule.estimated_qty_tonnes,
            schedule.corridor_name,
            schedule.notes,
        ])

    header_fill = PatternFill("solid", fgColor="173158")
    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = header_fill

    ws.freeze_panes = "A2"
    for column in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[get_column_letter(column[0].column)].width = min(max_len + 2, 32)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"scheduled-trucks-{status}-{today}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── POST /api/schedules/import ───────────────────────────────────────────────

_MATERIAL_TO_CORRIDOR = {
    "CLINKER":  "NORTHERN",
    "COAL":     "SOUTHERN_HIGHLANDS",
    "GYPSUM":   "SOUTHERN_COAST",
    "IRON ORE": "CENTRAL",
}

_HEADER_MAP = {
    # PO reference
    "po ref": "odoo_po_name", "po reference": "odoo_po_name", "po no": "odoo_po_name",
    "po number": "odoo_po_name",
    # material
    "material": "raw_material_type", "raw material": "raw_material_type",
    "material type": "raw_material_type", "product type": "raw_material_type",
    "aina ya bidhaa": "raw_material_type",
    # transporter
    "transporter": "transporter_name", "transporter name": "transporter_name",
    # driver
    "driver": "driver_name", "driver name": "driver_name",
    "jina la dereva": "driver_name", "jina la dereva (kama leseni)": "driver_name",
    # phone
    "phone": "driver_phone", "mobile": "driver_phone", "driver phone": "driver_phone",
    "driver mobile": "driver_phone", "phone number": "driver_phone",
    "contact": "driver_phone", "namba ya simu": "driver_phone",
    # licence
    "licence": "driver_license_no", "license": "driver_license_no",
    "driver licence": "driver_license_no", "driver license": "driver_license_no",
    "licence no": "driver_license_no", "license no": "driver_license_no",
    "driving licence": "driver_license_no", "driving license": "driver_license_no",
    "namba ya leseni": "driver_license_no", "licence no.": "driver_license_no",
    "license no.": "driver_license_no",
    # truck plate
    "truck no": "truck_plate", "truck no.": "truck_plate", "vehicle": "truck_plate",
    "plate": "truck_plate", "truck plate": "truck_plate", "truck number": "truck_plate",
    "namba za gari": "truck_plate",
    # trailer
    "trailer": "dealer_number", "trailer no": "dealer_number",
    "trailer no.": "dealer_number", "trailer number": "dealer_number",
    "namba za tela": "dealer_number",
    # origin/location
    "location": "origin_region", "origin": "origin_region", "origin region": "origin_region",
    "preferred location": "origin_region",
    # dispatch date
    "date of dispatch": "dispatch_date", "dispatch date": "dispatch_date",
    "dispatched": "dispatch_date", "date": "dispatch_date",
    # quantity
    "qty (mt)": "estimated_qty_tonnes", "qty mt": "estimated_qty_tonnes",
    "quantity": "estimated_qty_tonnes", "qty": "estimated_qty_tonnes",
    "tons": "estimated_qty_tonnes", "tonnes": "estimated_qty_tonnes",
    "loading tons": "estimated_qty_tonnes", "tonne": "estimated_qty_tonnes",
    "tani za kubeba": "estimated_qty_tonnes",
}


@router.post("/import")
async def import_schedules_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """
    Import truck schedules from an Excel file.
    File is processed entirely in memory — never written to disk.
    Duplicate rows (same truck_plate + dispatch_date) are skipped.
    Terminal records older than 30 days are auto-purged after import.
    """
    import uuid
    from datetime import timedelta

    from openpyxl import load_workbook

    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="File must be an Excel file (.xlsx or .xls)")

    # Read entirely into memory; never touch disk
    content = await file.read()
    await file.close()

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Cannot read Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # Build column index map from header row
    # Strip trailing /  .  '  and bracketed content e.g. "(Loading Tons) /"
    import re as _re
    def _clean_header(raw) -> str:
        if raw is None:
            return ""
        s = str(raw).strip().lower()
        s = _re.sub(r"[\s/.'\"()\-]+$", "", s)   # strip trailing punctuation
        s = _re.sub(r"^\s*\(.*?\)\s*", "", s)     # strip leading (...)
        return s.strip()

    header_row = [_clean_header(c) for c in rows[0]]
    col_map: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        field = _HEADER_MAP.get(h)
        if field and field not in col_map:
            col_map[field] = idx

    if "truck_plate" not in col_map:
        raise HTTPException(
            status_code=400,
            detail="Column 'Truck No.' not found. Check header row matches the template.",
        )

    now = datetime.now(timezone.utc)
    imported = 0
    skipped = 0
    errors: list[str] = []

    for row_num, row in enumerate(rows[1:], start=2):
        def cell(field: str):
            idx = col_map.get(field)
            return row[idx] if idx is not None and idx < len(row) else None

        plate_raw = cell("truck_plate")
        truck_plate = str(plate_raw).strip().upper() if plate_raw else None
        if not truck_plate:
            skipped += 1
            continue

        # Parse dispatch_date
        dispatch_date = None
        dd_raw = cell("dispatch_date")
        if dd_raw:
            if isinstance(dd_raw, datetime):
                dispatch_date = dd_raw.replace(tzinfo=timezone.utc) if dd_raw.tzinfo is None else dd_raw
            else:
                try:
                    from dateutil import parser as dateparser
                    dispatch_date = dateparser.parse(str(dd_raw)).replace(tzinfo=timezone.utc)
                except Exception:
                    errors.append(f"Row {row_num}: invalid dispatch date '{dd_raw}' — imported without date")

        # Dedup: skip if truck_plate + dispatch_date already exists
        dup_q = select(TruckSchedule.id).where(TruckSchedule.truck_plate == truck_plate)
        if dispatch_date:
            dup_q = dup_q.where(TruckSchedule.dispatch_date == dispatch_date)
        dup = await db.execute(dup_q.limit(1))
        if dup.scalar_one_or_none() is not None:
            skipped += 1
            continue

        # Resolve optional fields
        raw_material = str(cell("raw_material_type") or "").strip().upper() or None
        origin_region = str(cell("origin_region") or "").strip().upper() or "UNKNOWN"
        corridor = _MATERIAL_TO_CORRIDOR.get(raw_material, None) if raw_material else None

        qty_raw = cell("estimated_qty_tonnes")
        try:
            qty = float(qty_raw) if qty_raw is not None else 30.0
        except (ValueError, TypeError):
            qty = 30.0

        def _str(field: str) -> str | None:
            v = cell(field)
            return str(v).strip() if v is not None else None

        schedule = TruckSchedule(
            schedule_ref=f"IMP-{uuid.uuid4().hex[:10].upper()}",
            odoo_po_name=_str("odoo_po_name"),
            transporter_name=_str("transporter_name"),
            driver_name=_str("driver_name"),
            driver_phone=_str("driver_phone"),
            driver_license_no=_str("driver_license_no"),
            dealer_number=_str("dealer_number"),
            truck_plate=truck_plate,
            origin_region=origin_region,
            raw_material_type=raw_material,
            corridor_name=corridor,
            estimated_qty_tonnes=qty,
            dispatch_date=dispatch_date,
            upload_date=now,
            status=TruckScheduleStatus.EXPECTED,
            allocation_status=AllocationStatus.UNALLOCATED,
        )
        db.add(schedule)
        imported += 1

    await db.flush()

    # Auto-cleanup: delete terminal records older than 30 days
    cutoff = now - timedelta(days=30)
    cleanup_result = await db.execute(
        delete(TruckSchedule).where(
            TruckSchedule.allocation_status.in_(["LOADED", "RELEASED"]),
            TruckSchedule.upload_date < cutoff,
        ).returning(TruckSchedule.id)
    )
    cleaned_up = len(cleanup_result.fetchall())

    await db.commit()

    if imported > 0:
        broadcast_sse("schedules_imported", {"imported": imported})

    return {
        "imported": imported,
        "skipped": skipped,
        "cleaned_up": cleaned_up,
        "errors": errors,
    }


# ── GET /api/schedules/odoo-config ───────────────────────────────────────────

@router.get("/odoo-config")
async def get_odoo_config():
    """
    Returns Odoo SO creation URL config for client-side URL building.
    The Allocate button uses this to open a pre-filled Odoo SO form.
    """
    from app.config import get_settings
    s = get_settings()
    return {
        "odoo_url": s.odoo_url,
        "so_action_id": s.odoo_so_action_id,
        "so_menu_id": s.odoo_so_menu_id,
        "cids": s.odoo_so_cids,
        "fields": {
            "truck_no":       s.odoo_so_field_truck_no,
            "trailer_no":     s.odoo_so_field_trailer_no,
            "driver_name":    s.odoo_so_field_driver_name,
            "driver_phone":   s.odoo_so_field_driver_phone,
            "driver_license": s.odoo_so_field_driver_license,
        },
    }


# ── GET /api/schedules/{id}/odoo-url ─────────────────────────────────────────

@router.get("/{schedule_id}/odoo-url")
async def get_odoo_url(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """
    Build the Odoo SO creation URL for a specific truck schedule.
    Looks up fleet.vehicle by plate and driver.master.custom by license
    so the many2one fields are pre-filled (which triggers onchange to fill
    the readonly char fields like Truck No, Driver Name, etc.).
    """
    import asyncio as _aio
    from app.config import get_settings
    from app.services.odoo_sync import OdooClient

    schedule = await _get_or_404(schedule_id, db)
    s = get_settings()
    client = OdooClient()

    vehicle_odoo_id: int | None = None
    driver_odoo_id: int | None = None

    try:
        uid = await _aio.to_thread(client._uid_or_auth)
        models = client._models()

        if schedule.truck_plate:
            rows = await _aio.to_thread(
                models.execute_kw,
                s.odoo_db, uid, s.odoo_password,
                "fleet.vehicle", "search_read",
                [[["license_plate", "=", schedule.truck_plate]]],
                {"fields": ["id"], "limit": 1},
            )
            if rows:
                vehicle_odoo_id = rows[0]["id"]

        if schedule.driver_license_no:
            rows = await _aio.to_thread(
                models.execute_kw,
                s.odoo_db, uid, s.odoo_password,
                "driver.master.custom", "search_read",
                [[["license", "=", schedule.driver_license_no]]],
                {"fields": ["id"], "limit": 1},
            )
            if rows:
                driver_odoo_id = rows[0]["id"]
    except Exception:
        pass  # non-critical — open form without pre-fill if Odoo unreachable

    hash_parts = (
        f"cids={s.odoo_so_cids}"
        f"&menu_id={s.odoo_so_menu_id}"
        f"&action={s.odoo_so_action_id}"
        f"&model=sale.order&view_type=form"
    )
    if vehicle_odoo_id:
        hash_parts += f"&default_vehicle_id={vehicle_odoo_id}"
    if driver_odoo_id:
        hash_parts += f"&default_custom_driver_id={driver_odoo_id}"

    return {
        "url": f"{s.odoo_url}/web#{hash_parts}",
        "vehicle_found": vehicle_odoo_id is not None,
        "driver_found": driver_odoo_id is not None,
        "truck_plate": schedule.truck_plate,
        "driver_name": schedule.driver_name,
    }


# ── PATCH /api/schedules/{id}/mark-allocated — Option A ──────────────────────

@router.patch("/{schedule_id}/mark-allocated")
async def mark_allocated(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """
    Dispatcher manually marks a truck as allocated after completing SO in Odoo.
    Moves the truck from the available list to the Allocated section.
    """
    schedule = await _get_or_404(schedule_id, db)
    schedule.allocation_status = AllocationStatus.WAITING_LOADING
    await db.commit()
    broadcast_sse("schedule_updated", {
        "schedule_id": schedule.id,
        "allocation_status": schedule.allocation_status,
        "truck_plate": schedule.truck_plate,
    })
    return {"ok": True, "schedule_id": schedule_id, "allocation_status": schedule.allocation_status}


# ── POST /api/schedules/sync-allocations — Option B manual trigger ─────────────

@router.post("/sync-allocations")
async def sync_allocations(db: AsyncSession = Depends(get_db)):
    """
    Cross-reference Odoo sale orders (via REST API) against local truck schedules.
    Any truck whose plate appears in an Odoo SO is automatically marked as allocated.
    """
    import asyncio as _aio
    from app.services.odoo_sync import OdooClient

    client = OdooClient()
    so_rows = await _aio.to_thread(client.fetch_sale_orders_rest, 90)

    # Build set of truck plates that exist in Odoo SOs
    odoo_plates = {
        (r.get("truck_no") or "").replace(" ", "").upper()
        for r in so_rows
        if r.get("truck_no")
    }

    if not odoo_plates:
        return {"ok": True, "matched": 0, "message": "No truck plates found in Odoo SOs"}

    # Find unallocated schedules whose plate matches
    result = await db.execute(
        select(TruckSchedule).where(
            TruckSchedule.allocation_status == AllocationStatus.UNALLOCATED,
            TruckSchedule.truck_plate.isnot(None),
        )
    )
    schedules = result.scalars().all()

    matched = 0
    for s in schedules:
        plate = (s.truck_plate or "").replace(" ", "").upper()
        if plate in odoo_plates:
            s.allocation_status = AllocationStatus.WAITING_LOADING
            broadcast_sse("schedule_updated", {
                "schedule_id": s.id,
                "allocation_status": s.allocation_status,
                "truck_plate": s.truck_plate,
            })
            matched += 1

    if matched:
        await db.commit()

    return {"ok": True, "matched": matched, "odoo_plates_checked": len(odoo_plates)}


# ── GET /api/schedules/order-status ──────────────────────────────────────────

@router.get("/order-status")
async def get_order_status(db: AsyncSession = Depends(get_db)):
    """
    Returns all active truck schedules with their local allocation status.
    No Odoo sync — status reflects what is tracked in the local DB only.
    """
    q = select(TruckSchedule).where(
        TruckSchedule.status.notin_(["CANCELLED", "COMPLETED"])
    ).order_by(TruckSchedule.expected_arrival_dt.asc().nulls_last())
    result = await db.execute(q)
    schedules = result.scalars().all()

    return [
        {
            "schedule_id":        s.id,
            "schedule_ref":       s.schedule_ref,
            "expected_arrival_dt": s.expected_arrival_dt.isoformat() if s.expected_arrival_dt else None,
            "truck_plate":        s.truck_plate,
            "driver_name":        s.driver_name,
            "driver_license_no":  s.driver_license_no,
            "dealer_number":      s.dealer_number,
            "transporter_name":   s.transporter_name,
            "corridor_name":      s.corridor_name,
            "origin_region":      s.origin_region,
            "raw_material_type":  s.raw_material_type,
            "odoo_po_name":       s.odoo_po_name,
            "truck_status":       s.status,
            "allocation_status":  s.allocation_status,
        }
        for s in schedules
    ]


# ── POST /api/schedules/bulk-delete ──────────────────────────────────────────

@router.post("/bulk-delete")
async def bulk_delete_schedules(
    payload: dict,
    db: AsyncSession = Depends(get_db),
):
    """
    Delete multiple truck schedules by ID.
    Body: { "ids": [1, 2, 3] }
    """
    ids = payload.get("ids", [])
    if not ids:
        return {"ok": True, "deleted": 0}

    # Remove linked proposals first (FK constraint)
    await db.execute(
        delete(AllocationProposal).where(AllocationProposal.schedule_id.in_(ids))
    )

    result = await db.execute(
        delete(TruckSchedule)
        .where(TruckSchedule.id.in_(ids))
        .returning(TruckSchedule.id)
    )
    deleted = len(result.fetchall())
    await db.commit()

    if deleted:
        broadcast_sse("schedules_deleted", {"count": deleted})

    return {"ok": True, "deleted": deleted}


# ── DELETE /api/schedules/{id} ────────────────────────────────────────────────

@router.delete("/{schedule_id}")
async def delete_schedule(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """Delete a single truck schedule and its linked allocation proposals."""
    schedule = await _get_or_404(schedule_id, db)

    # Remove linked proposals first
    await db.execute(
        delete(AllocationProposal).where(AllocationProposal.schedule_id == schedule_id)
    )
    await db.delete(schedule)
    await db.commit()

    broadcast_sse("schedule_deleted", {"schedule_id": schedule_id})

    return {"ok": True, "deleted": schedule_id}


# ── GET /api/schedules/{id} ───────────────────────────────────────────────────

@router.get("/{schedule_id}", response_model=TruckScheduleRead)
async def get_schedule(schedule_id: int, db: AsyncSession = Depends(get_db)):
    schedule = await _get_or_404(schedule_id, db)
    return TruckScheduleRead.model_validate(schedule)


# ── PATCH /api/schedules/{id}/confirm-details ─────────────────────────────────

@router.patch("/{schedule_id}/confirm-details")
async def confirm_details(
    schedule_id: int,
    payload: TruckScheduleDetail,
    db: AsyncSession = Depends(get_db),
):
    """Add truck plate, driver info, and actual capacity (transporter pre-advice)."""
    schedule = await _get_or_404(schedule_id, db)

    if payload.truck_plate:
        schedule.truck_plate = payload.truck_plate
    if payload.driver_name:
        schedule.driver_name = payload.driver_name
    if payload.driver_phone:
        schedule.driver_phone = payload.driver_phone
    if payload.actual_capacity_tonnes:
        schedule.actual_capacity_tonnes = payload.actual_capacity_tonnes
    if payload.notes:
        schedule.notes = payload.notes

    # Promote to PRE_CONFIRMED if we now have truck details
    if schedule.truck_plate and schedule.status == TSS.EXPECTED:
        schedule.status = TSS.PRE_CONFIRMED

    await db.commit()

    # Broadcast SSE update
    broadcast_sse("schedule_updated", {
        "schedule_id": schedule_id,
        "schedule_ref": schedule.schedule_ref,
        "status": schedule.status,
        "truck_plate": schedule.truck_plate,
    })

    return {"ok": True, "schedule_ref": schedule.schedule_ref, "status": schedule.status}


# ── PATCH /api/schedules/{id}/arrived ────────────────────────────────────────

@router.patch("/{schedule_id}/arrived")
async def mark_arrived(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """Mark truck as physically arrived at Kimbiji Plant."""
    schedule = await _get_or_404(schedule_id, db)

    schedule.status = TSS.ARRIVED
    schedule.actual_arrival_dt = datetime.now(timezone.utc)
    await db.commit()

    broadcast_sse("truck_arrived", {
        "schedule_id": schedule_id,
        "schedule_ref": schedule.schedule_ref,
        "truck_plate": schedule.truck_plate,
        "origin_region": schedule.origin_region,
    })

    return {"ok": True, "schedule_ref": schedule.schedule_ref, "arrived_at": schedule.actual_arrival_dt}


# ── PATCH /api/schedules/{id}/dispatch ───────────────────────────────────────

@router.patch("/{schedule_id}/dispatch")
async def mark_dispatched(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """Mark truck as loaded and dispatched."""
    schedule = await _get_or_404(schedule_id, db)

    schedule.status = TSS.DISPATCHED
    schedule.dispatched_at = datetime.now(timezone.utc)
    schedule.allocation_status = AllocationStatus.DISPATCHED

    result = await db.execute(
        select(AllocationProposal).where(
            AllocationProposal.schedule_id == schedule_id,
            AllocationProposal.status == "CONFIRMED",
        ).limit(1)
    )
    proposal = result.scalar_one_or_none()
    if proposal:
        proposal.dispatched_at = datetime.now(timezone.utc)
        proposal.status = "DISPATCHED"

    await db.commit()

    broadcast_sse("truck_dispatched", {
        "schedule_id": schedule_id,
        "schedule_ref": schedule.schedule_ref,
        "truck_plate": schedule.truck_plate,
    })

    return {"ok": True, "schedule_ref": schedule.schedule_ref}


# ── POST /api/schedules/{id}/rematch ─────────────────────────────────────────

@router.post("/{schedule_id}/rematch")
async def force_rematch(schedule_id: int, db: AsyncSession = Depends(get_db)):
    """Force re-run of the matching engine for this schedule."""
    schedule = await _get_or_404(schedule_id, db)

    from app.services.matching_engine import MatchingEngine
    from app.models.matching_event import MatchTrigger

    engine = MatchingEngine(db)
    proposals = await engine.rematch(schedule, trigger=MatchTrigger.MANUAL)
    await db.commit()

    broadcast_sse("proposals_updated", {
        "schedule_id": schedule_id,
        "schedule_ref": schedule.schedule_ref,
        "proposals_count": len(proposals),
    })

    return {
        "ok": True,
        "schedule_ref": schedule.schedule_ref,
        "proposals_generated": len(proposals),
    }


# ── GET /api/schedules/feed — SSE ────────────────────────────────────────────

@router.get("/feed")
async def sse_feed():
    """
    Server-Sent Events stream for live dashboard updates.
    The browser connects once and receives push events as trucks are
    allocated, arrive, or proposals are confirmed.
    """
    import uuid
    client_id = str(uuid.uuid4())
    queue: asyncio.Queue = asyncio.Queue(maxsize=50)
    _sse_clients[client_id] = queue

    async def event_generator() -> AsyncGenerator[str, None]:
        # Send initial connection confirmation
        yield f"data: {json.dumps({'type': 'connected', 'client_id': client_id})}\n\n"
        try:
            while True:
                try:
                    payload = await asyncio.wait_for(queue.get(), timeout=30.0)
                    yield f"data: {payload}\n\n"
                except asyncio.TimeoutError:
                    # Heartbeat — keep connection alive
                    yield f"data: {json.dumps({'type': 'heartbeat'})}\n\n"
        except asyncio.CancelledError:
            pass
        finally:
            _sse_clients.pop(client_id, None)

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── Helpers ───────────────────────────────────────────────────────────────────

async def _get_or_404(schedule_id: int, db: AsyncSession) -> TruckSchedule:
    result = await db.execute(
        select(TruckSchedule).where(TruckSchedule.id == schedule_id)
    )
    schedule = result.scalar_one_or_none()
    if not schedule:
        raise HTTPException(status_code=404, detail=f"Schedule {schedule_id} not found")
    return schedule
