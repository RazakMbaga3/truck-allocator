"""
app/routers/orders.py — CementOrder API endpoints.

GET /api/orders                     — All synced orders
GET /api/orders/unallocated         — Orders awaiting a truck
GET /api/orders/near-ready          — Near-ready (not dispatch_ready but eligible by ETA)
GET /api/orders/by-corridor/{name}  — Orders filtered by corridor
POST /api/orders/sync               — Force full Odoo sync
"""

from __future__ import annotations

from fastapi import APIRouter, Depends, Query
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.database import get_db
from app.models import CementOrder, OrderAllocationStatus
from app.schemas.cement_order import CementOrderListItem, CementOrderRead

router = APIRouter(prefix="/api/orders", tags=["orders"])


@router.get("", response_model=list[CementOrderListItem])
async def list_orders(
    limit: int = Query(200, ge=1, le=1000),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(CementOrder)
        .order_by(CementOrder.deadline_dt.asc().nulls_last())
        .limit(limit)
    )
    return [CementOrderListItem.model_validate(o) for o in result.scalars().all()]


@router.get("/unallocated", response_model=list[CementOrderListItem])
async def list_unallocated(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(CementOrder)
        .where(
            CementOrder.allocation_status == OrderAllocationStatus.UNALLOCATED,
            CementOrder.return_load_eligible == True,
        )
        .order_by(CementOrder.urgency_score.desc())
    )
    return [CementOrderListItem.model_validate(o) for o in result.scalars().all()]


@router.get("/near-ready", response_model=list[CementOrderListItem])
async def list_near_ready(db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(CementOrder)
        .where(
            CementOrder.near_ready == True,
            CementOrder.allocation_status == OrderAllocationStatus.UNALLOCATED,
        )
        .order_by(CementOrder.near_ready_eta.asc().nulls_last())
    )
    return [CementOrderListItem.model_validate(o) for o in result.scalars().all()]


@router.get("/by-corridor/{corridor}", response_model=list[CementOrderListItem])
async def list_by_corridor(corridor: str, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(CementOrder)
        .where(
            CementOrder.delivery_corridor == corridor.upper(),
            CementOrder.allocation_status == OrderAllocationStatus.UNALLOCATED,
        )
        .order_by(CementOrder.urgency_score.desc())
    )
    return [CementOrderListItem.model_validate(o) for o in result.scalars().all()]


# ── GET /api/orders/live-status — Order Status page ──────────────────────────

@router.get("/live-status")
async def get_live_status(days: int = Query(7, ge=1, le=90)):
    """
    Fetch recent Sale Orders from Odoo REST API for the Order Status page.
    Returns: date, customer, location, transporter, driver, license, qty, status.
    """
    import asyncio as _aio
    from app.services.odoo_sync import OdooClient

    client = OdooClient()
    rows = await _aio.to_thread(client.fetch_sale_orders_rest, days)

    STATE_MAP = {
        "draft": "Pending", "sent": "Pending", "sale": "Pending",
        "done": "Dispatched", "cancel": "Cancelled",
    }

    result = []
    for r in rows:
        driver      = r.get("driver_name") or None
        location    = r.get("destination_location") or None
        transporter = r.get("transporter_name") or None
        # Skip SOs with no truck/driver details — not actionable for dispatch
        if not driver and not transporter and not location:
            continue

        # delivery_status is the authoritative dispatch field
        delivery_status = r.get("delivery_status", "")
        if delivery_status == "delivered":
            status = "Dispatched"
        else:
            status = "Pending"

        # qty_mt: UOM varies — "50 KG BAG" needs ÷20, "MT" is already in tonnes
        qty_mt = 0.0
        for line in (r.get("order_lines") or []):
            qty = float(line.get("product_uom_qty", 0))
            if line.get("uom") == "MT":
                qty_mt += qty
            else:
                qty_mt += qty / 20.0

        result.append({
            "so_name":      r.get("name"),
            "date_order":   r.get("date_order"),
            "customer":     r.get("partner_name"),
            "location":     location,
            "transporter":  transporter,
            "driver":       driver,
            "license":      r.get("driver_license") or None,
            "driver_phone": r.get("driver_mobile") or None,
            "qty_mt":       qty_mt,
            "status":       status,
        })
    return result


# ── GET /api/orders/live-status/export — Excel download ──────────────────────

@router.get("/live-status/export")
async def export_live_status(
    days: int = Query(7, ge=1, le=90),
    status: str = Query("", alias="status"),
):
    """Export Order Status data as branded Excel file."""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    rows = await get_live_status(days=days)
    if status:
        rows = [r for r in rows if r.get("status") == status]

    wb = Workbook()
    ws = wb.active
    ws.title = "Order Status"

    NAVY, ORANGE, WHITE, GRAY = "173158", "F49545", "FFFFFF", "F5F5F5"
    hdr_font   = Font(bold=True, color=WHITE, size=10)
    hdr_fill   = PatternFill("solid", fgColor=NAVY)
    hdr_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin       = Side(style="thin", color="D3D3D3")
    border     = Border(left=thin, right=thin, top=thin, bottom=thin)
    center     = Alignment(horizontal="center", vertical="center")
    left       = Alignment(horizontal="left",   vertical="center")
    right      = Alignment(horizontal="right",  vertical="center")

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value = "NYATI CEMENT — ORDER STATUS REPORT"
    c.font  = Font(bold=True, color=WHITE, size=13)
    c.fill  = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:J2")
    s = ws["A2"]
    s.value = f"Period: Last {days} days  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    s.font  = Font(color=ORANGE, size=9)
    s.fill  = PatternFill("solid", fgColor="EEF2F7")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    headers = ["SO No.", "Date", "Customer", "Location", "Transporter",
               "Driver", "Phone", "Licence", "Qty (MT)", "Status"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = hdr_align; cell.border = border
    ws.row_dimensions[3].height = 20

    def fmt_date(dt):
        if not dt: return ""
        try: return str(dt)[:10]
        except: return str(dt)

    aligns = [center, center, left, left, left, left, center, center, right, center]
    for i, r in enumerate(rows):
        rn   = i + 4
        fill = PatternFill("solid", fgColor=GRAY if i % 2 == 0 else WHITE)
        vals = [
            r.get("so_name") or "",
            fmt_date(r.get("date_order")),
            r.get("customer") or "",
            r.get("location") or "",
            r.get("transporter") or "",
            r.get("driver") or "",
            r.get("driver_phone") or "",
            r.get("license") or "",
            r.get("qty_mt") or 0,
            r.get("status") or "",
        ]
        for col, (val, aln) in enumerate(zip(vals, aligns), 1):
            cell = ws.cell(row=rn, column=col, value=val)
            cell.fill = fill; cell.alignment = aln
            cell.border = border; cell.font = Font(size=9)
            if col == 10:
                color = NAVY if val == "Dispatched" else ORANGE
                cell.font = Font(size=9, bold=True, color=color)
        ws.row_dimensions[rn].height = 15

    for col, w in enumerate([14,12,22,18,20,18,14,14,10,12], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"order-status-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ── GET /api/orders/final-status — Final Status page ─────────────────────────

@router.get("/final-status")
async def get_final_status(
    days: int = Query(30, ge=1, le=90),
    db: AsyncSession = Depends(get_db),
):
    """
    Fetch SOs from Odoo REST API + PO reference from local TruckSchedule.
    Returns: po_ref, so_name, transporter, driver, location, qty,
             invoice_no, invoice_date, status (Dispatched/Released), remark.
    """
    import asyncio as _aio
    from sqlalchemy import select as sa_select
    from app.models import TruckSchedule
    from app.services.odoo_sync import OdooClient

    client = OdooClient()
    so_rows = await _aio.to_thread(client.fetch_sale_orders_rest, days)

    # Filter out cancelled and draft — exclude orders with no operational data
    so_rows = [r for r in so_rows if r.get("state") not in ("cancel", "draft")]

    # Look up PO references from local DB by truck plate
    plates = [r["truck_no"] for r in so_rows if r.get("truck_no")]
    po_by_plate: dict[str, str] = {}
    if plates:
        res = await db.execute(
            sa_select(TruckSchedule.truck_plate, TruckSchedule.odoo_po_name).where(
                TruckSchedule.truck_plate.in_(plates),
                TruckSchedule.odoo_po_name.isnot(None),
            )
        )
        po_by_plate = {row.truck_plate: row.odoo_po_name for row in res}

    STATE_MAP = {"done": "Dispatched", "sale": "Pending", "sent": "Pending"}

    output = []
    for r in so_rows:
        driver      = r.get("driver_name") or None
        location    = r.get("destination_location") or None
        transporter = r.get("transporter_name") or None

        if not driver and not location and not transporter:
            continue

        # delivery_status + invoice_status determine final outcome
        delivery_status = r.get("delivery_status", "")
        invoice_status  = r.get("invoice_status", "")
        if delivery_status == "delivered" and invoice_status == "invoiced":
            status = "Dispatched"
        elif delivery_status == "delivered":
            status = "Dispatched"
        else:
            status = "Released"

        # qty_mt: UOM varies — "50 KG BAG" needs ÷20, "MT" is already in tonnes
        qty_mt = 0.0
        for line in (r.get("order_lines") or []):
            qty = float(line.get("product_uom_qty", 0))
            if line.get("uom") == "MT":
                qty_mt += qty
            else:
                qty_mt += qty / 20.0

        truck_no = r.get("truck_no") or ""
        output.append({
            "po_ref":       po_by_plate.get(truck_no),
            "so_name":      r.get("name"),
            "transporter":  transporter,
            "driver":       driver,
            "license":      r.get("driver_license") or None,
            "location":     location,
            "qty_mt":       qty_mt,
            "invoice_no":   r.get("invoice_no") or None,
            "invoice_date": r.get("invoice_date") or None,
            "status":       status,
            "remark":       r.get("note") or "",
            "date_order":   r.get("date_order"),
        })
    return output


# ── GET /api/orders/final-status/export — Excel download ─────────────────────

@router.get("/final-status/export")
async def export_final_status(
    days: int = Query(30, ge=1, le=90),
    status: str = Query("", alias="status"),
    db: AsyncSession = Depends(get_db),
):
    """Export Final Status data as branded Excel file."""
    from fastapi.responses import StreamingResponse
    from io import BytesIO
    from datetime import datetime, timedelta, timezone
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import asyncio as _aio
    from app.services.odoo_sync import OdooClient
    from app.config import get_settings

    # Reuse the same data fetch as get_final_status
    rows = (await get_final_status(days=days, db=db))

    if status:
        rows = [r for r in rows if r.get("status") == status]

    # Build Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Final Status"

    NAVY   = "173158"
    ORANGE = "F49545"
    WHITE  = "FFFFFF"
    GRAY   = "F5F5F5"

    header_font    = Font(bold=True, color=WHITE, size=10)
    header_fill    = PatternFill("solid", fgColor=NAVY)
    header_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_side      = Side(style="thin", color="D3D3D3")
    thin_border    = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align   = Alignment(horizontal="center", vertical="center")
    left_align     = Alignment(horizontal="left",   vertical="center")
    right_align    = Alignment(horizontal="right",  vertical="center")

    # Title row
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = f"NYATI CEMENT — FINAL STATUS REPORT"
    title_cell.font  = Font(bold=True, color=WHITE, size=13)
    title_cell.fill  = PatternFill("solid", fgColor=NAVY)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells("A2:J2")
    sub_cell = ws["A2"]
    sub_cell.value = f"Period: Last {days} days  |  Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    sub_cell.font  = Font(color=ORANGE, size=9)
    sub_cell.fill  = PatternFill("solid", fgColor="EEF2F7")
    sub_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    # Headers
    headers = ["PO Ref", "SO No.", "Transporter", "Driver", "Location",
               "Qty (MT)", "Invoice No.", "Invoice Date", "Status", "Remark"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align
        cell.border    = thin_border
    ws.row_dimensions[3].height = 20

    # Data rows
    for i, r in enumerate(rows):
        row_num = i + 4
        fill = PatternFill("solid", fgColor=GRAY if i % 2 == 0 else WHITE)
        values = [
            r.get("po_ref") or "",
            r.get("so_name") or "",
            r.get("transporter") or "",
            r.get("driver") or "",
            r.get("location") or "",
            r.get("qty_mt") or 0,
            r.get("invoice_no") or "",
            r.get("invoice_date") or "",
            r.get("status") or "",
            r.get("remark") or "",
        ]
        aligns = [center_align, center_align, left_align, left_align, left_align,
                  right_align, center_align, center_align, center_align, left_align]
        for col, (val, aln) in enumerate(zip(values, aligns), 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.fill      = fill
            cell.alignment = aln
            cell.border    = thin_border
            cell.font      = Font(size=9)
            if col == 9:  # Status column — color by value
                if val == "Released":
                    cell.font = Font(size=9, bold=True, color=ORANGE)
                elif val == "Dispatched":
                    cell.font = Font(size=9, bold=True, color=NAVY)
        ws.row_dimensions[row_num].height = 15

    # Column widths
    col_widths = [14, 16, 22, 18, 20, 10, 16, 14, 12, 30]
    for col, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A4"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"final-status-{datetime.now().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── GET /api/orders/{order_id} ────────────────────────────────────────────────

@router.get("/{order_id}", response_model=CementOrderRead)
async def get_order(order_id: int, db: AsyncSession = Depends(get_db)):
    from fastapi import HTTPException
    result = await db.execute(select(CementOrder).where(CementOrder.id == order_id))
    order = result.scalar_one_or_none()
    if not order:
        raise HTTPException(status_code=404, detail=f"Order {order_id} not found")
    return CementOrderRead.model_validate(order)


@router.post("/sync")
async def sync_orders(db: AsyncSession = Depends(get_db)):
    """Force a full Odoo sync of purchase orders + sale orders."""
    from app.services.odoo_sync import OdooSyncService
    svc = OdooSyncService(db)
    stats = await svc.run_full_sync()
    return {"ok": True, "stats": stats}
